import streamlit as st
import pandas as pd
from openpyxl import Workbook
import zipfile
import struct
import io

# -----------------------------------------------------------------------------
# GCD Extraction Logic (Adapted from extract_gcd_data.py)
# -----------------------------------------------------------------------------
COL_CYCLENO   = 4
COL_STEPNO    = 6
COL_STEPTIME  = 7
COL_VOLTAGE   = 9

def _iter_records(raw: bytes):
    i = 0
    n = len(raw)
    while i < n:
        b0 = raw[i]; i += 1
        if b0 & 0x80:
            if i >= n: break
            b1 = raw[i]; i += 1
            rec_type = (b0 & 0x7F) | (b1 << 7)
        else:
            rec_type = b0

        size = 0; shift = 0
        for _ in range(4):
            if i >= n: break
            b = raw[i]; i += 1
            size |= (b & 0x7F) << shift
            shift += 7
            if not (b & 0x80): break

        rec_data = raw[i:i + size]
        i += size
        yield rec_type, rec_data

def _decode_xlwidestring(rec_data: bytes, str_offset: int = 8) -> str:
    if len(rec_data) < str_offset + 5:
        return ''
    cch   = struct.unpack_from('<I', rec_data, str_offset)[0]
    high  = rec_data[str_offset + 4]
    start = str_offset + 5
    nb    = cch * (2 if high else 1)
    raw   = rec_data[start: start + nb]
    enc   = 'utf-16-le' if high else 'latin-1'
    return raw.decode(enc, errors='replace').replace('\x00', '').strip()

def _parse_time(s: str) -> float:
    s = s.strip()
    if ':' in s:
        m, sec = s.split(':', 1)
        return float(m) * 60.0 + float(sec)
    return float(s)

def read_sheet(file_obj, sheet_entry: str = "xl/worksheets/sheet2.bin") -> dict:
    with zipfile.ZipFile(file_obj) as z:
        raw = z.read(sheet_entry)

    cycle, step, step_time, voltage = {}, {}, {}, {}
    cur_row = -1
    for rec_type, rec_data in _iter_records(raw):
        if rec_type == 0 and len(rec_data) >= 4:
            cur_row = struct.unpack_from('<I', rec_data, 0)[0]
            continue
        if cur_row <= 0:
            continue
        if rec_type == 62 and len(rec_data) >= 13:
            col = struct.unpack_from('<I', rec_data, 0)[0]
            if col == COL_STEPTIME:
                s = _decode_xlwidestring(rec_data, str_offset=8)
                if s:
                    try:
                        step_time[cur_row] = _parse_time(s)
                    except ValueError:
                        pass
        elif rec_type == 5 and len(rec_data) >= 16:
            col = struct.unpack_from('<I', rec_data, 0)[0]
            if col == COL_CYCLENO:
                cycle[cur_row] = struct.unpack_from('<d', rec_data, 8)[0]
            elif col == COL_STEPNO:
                step[cur_row] = struct.unpack_from('<d', rec_data, 8)[0]
            elif col == COL_VOLTAGE:
                voltage[cur_row] = struct.unpack_from('<d', rec_data, 8)[0]

    return {'cycle': cycle, 'step': step, 'step_time': step_time, 'voltage': voltage}

def process_gcd_file(file_obj) -> pd.DataFrame:
    data = read_sheet(file_obj)
    cycle, step, step_time, voltage = data['cycle'], data['step'], data['step_time'], data['voltage']

    rows_22, rows_31 = [], []
    all_rows = sorted(set(cycle) & set(step) & set(step_time) & set(voltage))
    for r in all_rows:
        c, s = cycle[r], step[r]
        if abs(c - 2) < 0.5 and abs(s - 2) < 0.5:
            rows_22.append(r)
        elif abs(c - 3) < 0.5 and abs(s - 1) < 0.5:
            rows_31.append(r)

    if not rows_22 and not rows_31:
        return None

    rows_22.sort()
    rows_31.sort()
    pieces = []
    offset = 0.0

    if rows_22:
        times = [step_time[r] for r in rows_22]
        volts = [voltage[r] for r in rows_22]
        t0 = times[0]
        elapsed = [t - t0 + offset for t in times]
        cadence = times[-1] - times[-2] if len(times) > 1 else 1.0
        offset = elapsed[-1] + abs(cadence)
        pieces.append(pd.DataFrame({'StepTime_s': elapsed, 'Voltage_V': volts}))

    if rows_31:
        times = [step_time[r] for r in rows_31]
        volts = [voltage[r] for r in rows_31]
        t0 = times[0]
        elapsed = [t - t0 + offset for t in times]
        pieces.append(pd.DataFrame({'StepTime_s': elapsed, 'Voltage_V': volts}))

    return pd.concat(pieces, ignore_index=True) if pieces else None

# -----------------------------------------------------------------------------
# CV Extraction Logic (Adapted from extract_CV_xlsb.py)
# -----------------------------------------------------------------------------
def process_cv_file(file_obj) -> pd.DataFrame:
    df = pd.read_excel(file_obj, sheet_name="DCData1", engine="pyxlsb", header=None)
    first_row = df.iloc[0].astype(str).tolist()
    has_headers = False
    for val in first_row:
        if "CycleNo" in str(val):
            has_headers = True
            break
            
    cycle_idx, curr_idx, volt_idx = 4, 8, 9
    
    if has_headers:
        cycle_idx = next((i for i, x in enumerate(first_row) if "CycleNo" in str(x)), 4)
        curr_idx = next((i for i, x in enumerate(first_row) if "_Current_A" in str(x)), 8)
        volt_idx = next((i for i, x in enumerate(first_row) if "Voltage_V" in str(x)), 9)
        df = df.iloc[1:].reset_index(drop=True)
    else:
        if df.iloc[0].isna().all():
            df = df.iloc[1:].reset_index(drop=True)

    cycle_col = df.iloc[:, cycle_idx]
    cycle_values = pd.to_numeric(cycle_col, errors='coerce').dropna()
    unique_cycles = sorted(cycle_values.unique())
    
    if len(unique_cycles) < 2:
        return None
        
    second_last_cycle = unique_cycles[-2]
    mask = (pd.to_numeric(cycle_col, errors='coerce') == second_last_cycle)
    filtered_df = df[mask]
    
    extracted = filtered_df.iloc[:, [volt_idx, curr_idx]].reset_index(drop=True)
    extracted.columns = ['Voltage_V', '_Current_A']
    return extracted

# -----------------------------------------------------------------------------
# Streamlit UI
# -----------------------------------------------------------------------------
st.set_page_config(page_title="Electrochemistry Data Extractor", layout="wide")
st.title("Electrochemistry Data Extractor")
st.markdown("Upload your `.xlsb` files and select the extraction method (CV or GCD).")

option = st.radio("Select Processing Mode:", ("CV", "GCD"))
uploaded_files = st.file_uploader("Upload .xlsb files", type=["xlsb"], accept_multiple_files=True)

if st.button("Process Files"):
    if not uploaded_files:
        st.warning("Please upload at least one .xlsb file.")
    else:
        results = []
        progress_bar = st.progress(0)
        
        for i, file in enumerate(uploaded_files):
            try:
                if option == "CV":
                    df = process_cv_file(file)
                else:
                    df = process_gcd_file(file)
                    
                if df is not None and not df.empty:
                    results.append((file.name, df))
                else:
                    st.warning(f"No valid data extracted from {file.name}.")
            except Exception as e:
                st.error(f"Failed to process {file.name}: {e}")
            progress_bar.progress((i + 1) / len(uploaded_files))
            
        if results:
            st.success(f"Successfully processed {len(results)} files. Preparing Excel output...")
            
            output = io.BytesIO()
            if option == "CV":
                final_df = pd.concat([res[1] for res in results], axis=1)
                wb = Workbook()
                ws = wb.active
                ws.title = "Combined Data"
                
                filenames = [res[0] for res in results]
                row1 = []
                for fn in filenames:
                    row1.extend([fn, ""])
                ws.append(row1)
                
                row2 = []
                for _ in filenames:
                    row2.extend(["Voltage_V", "_Current_A"])
                ws.append(row2)
                
                final_df.fillna("", inplace=True)
                for r in final_df.itertuples(index=False, name=None):
                    ws.append(r)
                wb.save(output)
            else:
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    wb = writer.book
                    ws = wb.create_sheet("Extracted Data")
                    col_start = 1
                    for file_name, df_data in results:
                        ws.cell(row=1, column=col_start, value=file_name)
                        ws.cell(row=2, column=col_start, value="StepTime_s")
                        ws.cell(row=2, column=col_start + 1, value="Voltage_V")
                        
                        for row_idx, (st_val, vv) in enumerate(zip(df_data["StepTime_s"], df_data["Voltage_V"]), start=3):
                            cell_t = ws.cell(row=row_idx, column=col_start, value=round(float(st_val), 4))
                            cell_t.number_format = "0.0000"
                            ws.cell(row=row_idx, column=col_start + 1, value=round(float(vv), 6))
                        col_start += 3
                        
                    if "Sheet" in wb.sheetnames:
                        del wb["Sheet"]
            
            output.seek(0)
            file_name = f"Combined_{option}_Data.xlsx"
            st.download_button(label="Download Combined Excel File", data=output, file_name=file_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.error("No valid data was extracted. Please check your files.")
